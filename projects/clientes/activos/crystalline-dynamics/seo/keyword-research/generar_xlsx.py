#!/usr/bin/env python3
"""Genera Excel .xlsx con 10 tabs formateados (colores MM Agency + Crystalline)."""
import csv, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
CSV_PATH = os.path.join(HERE, "keywords.csv")
OUT = os.path.join(HERE, "ocala-fence-keywords.xlsx")

# Paleta (brand Crystalline + MM Agency)
GREEN = "136229"
GREEN_LIGHT = "E8F5EB"
GOLD = "D1B487"
GOLD_LIGHT = "FDF6E3"
DARK = "1A1A1A"
GREY = "666666"
WHITE = "FFFFFF"
ORANGE = "D97706"
ORANGE_LIGHT = "FFF3E0"
RED = "991B1B"
RED_LIGHT = "FDE7E7"
BLUE = "1E40AF"
BLUE_LIGHT = "DBEAFE"
PURPLE = "4C1D95"
PURPLE_LIGHT = "EDE9FE"
GREY_LIGHT = "F1F5F9"

CLUSTERS_ORDER = [
    ("resumen", "🎯 Resumen Ejecutivo", GREEN, GREEN_LIGHT),
    ("money", "💰 Money Keywords", GREEN, GREEN_LIGHT),
    ("near-me", "📍 Near Me · GBP", GREEN, GREEN_LIGHT),
    ("product", "🏷️ Productos", GOLD, GOLD_LIGHT),
    ("product-niche", "⭐ Durafence (nicho)", GOLD, GOLD_LIGHT),
    ("geo-zip", "📌 ZIPs de Ocala", ORANGE, ORANGE_LIGHT),
    ("geo-county", "🗺️ Marion County", ORANGE, ORANGE_LIGHT),
    ("nearby-city", "🚗 Ciudades vecinas", ORANGE, ORANGE_LIGHT),
    ("financing", "💵 Financiamiento $0 Down", RED, RED_LIGHT),
    ("hispanic", "🇪🇸 Mercado Hispano", RED, RED_LIGHT),
    ("service", "⚡ Servicios + CTAs", BLUE, BLUE_LIGHT),
    ("blog", "📝 Blog / Content SEO", PURPLE, PURPLE_LIGHT),
    ("negative", "⛔ Negative Keywords (Ads)", GREY, GREY_LIGHT),
    ("todas", "📋 Todas las keywords", DARK, GREY_LIGHT),
]

CLUSTER_DESC = {
    "money": "Money keywords de Ocala + FL. Alto intent transaccional — cliente listo para comprar. Van en homepage (H1, meta, intro).",
    "near-me": "Búsquedas 'near me' — se ganan con Google Business Profile optimizado + proximity + reviews. Mucho volumen pero CPC alto si van por ads.",
    "product": "Una landing dedicada por tipo de producto: Vinyl, Aluminum, Privacy, Pool, Residential. Captura intent específico de compra.",
    "product-niche": "Durafence es producto de nicho con poca competencia SEO — oportunidad de capturar 100% del SERP con landing propia.",
    "geo-zip": "SEO programático: una página por ZIP (34471-34480) con contenido local único. Bajo volumen por página pero altísima conversión.",
    "geo-county": "Marion County es el condado donde está Ocala. Capturar búsquedas a nivel condado.",
    "nearby-city": "Expansión geográfica. The Villages (al sur) es el mercado grande con 90 vol/mes — oportunidad P1.",
    "financing": "USP principal del cliente (Aqua, Launch, Synchrony, Klarna, Affirm). Landing dedicada con CTA al aplicador.",
    "hispanic": "USP del cliente: habla español. Mercado hispano en Marion County sin competencia directa en español.",
    "service": "Quick-wins con 'free estimate' y 'same-day'. Van como CTAs en todas las páginas.",
    "blog": "Content SEO. Pillar: 'how much does a fence cost in florida' + 'fence for bears florida' (ángulo Ocala único).",
    "negative": "⚠️ Lista para pegar DIRECTO en Google Ads como 'negative keywords'. Excluye wood, chain link, repair, DIY, jobs, rental, used.",
}

def load():
    with open(CSV_PATH) as f:
        return list(csv.DictReader(f))

def border_thin(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def build_header(ws, title, subtitle, color, bg_color, cols=9):
    # Row 1: title mega (merge)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", size=22, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 44

    # Row 2: subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name="Calibri", size=11, italic=True, color=GREY)
    c.fill = PatternFill("solid", fgColor=bg_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 34

    # Row 3: spacer
    ws.row_dimensions[3].height = 8

def write_table(ws, rows, headers, color, bg_color, start_row=4):
    # Headers
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border_thin("AAAAAA")
    ws.row_dimensions[start_row].height = 28

    # Data rows
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=10, color=DARK)
            # Alternating bg
            if r_idx % 2 == 0:
                c.fill = PatternFill("solid", fgColor=bg_color)
            else:
                c.fill = PatternFill("solid", fgColor=WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            c.border = border_thin("EEEEEE")
        ws.row_dimensions[r_idx].height = 22

def kw_rows(items):
    """Convierte dicts a tuplas para tabla de keywords."""
    out = []
    for r in sorted(items, key=lambda x: (prio_sort(x["priority"]), -int(x["monthly_volume"] or 0))):
        out.append((
            r["keyword"],
            int(r["monthly_volume"]) if r["monthly_volume"].isdigit() else 0,
            f"${float(r['cpc_usd']):.2f}" if r["cpc_usd"] else "—",
            r["competition"].upper() if r["competition"] else "—",
            r["intent"].capitalize() if r["intent"] else "—",
            r["priority"],
            r["target_page"] or "—",
            r["notes"]
        ))
    return out

def prio_sort(p):
    return {"P1": 1, "P2": 2, "P3": 3}.get(p, 9)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, row=5, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)

def build():
    rows = load()
    wb = Workbook()
    # Remove default sheet
    del wb[wb.sheetnames[0]]

    total = len(rows)
    total_vol = sum(int(r["monthly_volume"]) for r in rows if r["monthly_volume"].isdigit())
    p1_count = sum(1 for r in rows if r["priority"] == "P1")
    p2_count = sum(1 for r in rows if r["priority"] == "P2")
    p3_count = sum(1 for r in rows if r["priority"] == "P3")
    avg_cpc = sum(float(r["cpc_usd"]) for r in rows) / total

    by_cluster = defaultdict(list)
    for r in rows:
        by_cluster[r["cluster"]].append(r)

    # =========================================================================
    # TAB 1 — RESUMEN EJECUTIVO
    # =========================================================================
    ws = wb.create_sheet("🎯 Resumen")
    build_header(ws, "🎯 Resumen Ejecutivo",
                 "Ocala Fence Install · Keyword Research · MM Agency × Crystalline Dynamics",
                 GREEN, GREEN_LIGHT, cols=6)
    # Stats row
    stats = [
        ("KEYWORDS TOTALES", f"{total}"),
        ("VOL/MES TOTAL", f"{total_vol:,}"),
        ("PRIORIDAD 1", f"{p1_count}"),
        ("PRIORIDAD 2", f"{p2_count}"),
        ("PRIORIDAD 3", f"{p3_count}"),
        ("CPC PROMEDIO", f"${avg_cpc:.2f}"),
    ]
    for i, (label, val) in enumerate(stats, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = Font(name="Calibri", size=9, bold=True, color=GREY)
        c.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin("AAAAAA")
        c = ws.cell(row=5, column=i, value=val)
        c.font = Font(name="Calibri", size=18, bold=True, color=GREEN)
        c.fill = PatternFill("solid", fgColor=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin("AAAAAA")
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 44

    # Estrategia
    ws.cell(row=7, column=1, value="📊 ESTRATEGIA PROPUESTA").font = Font(size=14, bold=True, color=GREEN)
    ws.row_dimensions[7].height = 30
    strategy = [
        "1. Homepage optimizada para 'fence company ocala fl' + 'fence contractor ocala fl' (money kw #1-#2, 530 vol/mes combinados)",
        "2. Crear 5 service pages (productos): Vinyl, Aluminum, Privacy, Pool, Residential — una landing dedicada por cada una",
        "3. Landing $0 Down Financing — USP diferenciador que captura búsquedas que la competencia no trabaja",
        "4. 6 páginas por ZIP (34471-34480) — SEO programático para capturar intent hiperlocal",
        "5. Landing dedicada para The Villages (90 vol/mes) — mercado grande al sur",
        "6. Landing en español para mercado hispano sin competencia directa en Marion County",
        "7. Blog hub con pillar 'how much does a fence cost in florida' (590 vol/mes) + 6 artículos satelitales",
        "8. Negative keywords excluidas en Google Ads: wood, chain link, repair, DIY, jobs, rental, used — ahorra ~30% del presupuesto",
        "",
        "⭐ ÁNGULO ÚNICO: 'fence for bears florida' (90 vol/mes) — ningún competidor tiene este artículo. Gana el 100% del SERP.",
    ]
    for i, line in enumerate(strategy, 8):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(size=11, color=DARK, bold=(i == 8+len(strategy)-1))
        c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 32

    # Top 10 keywords
    ws.cell(row=20, column=1, value="🏆 TOP 10 KEYWORDS POR VOLUMEN (TRANSACCIONALES)").font = Font(size=14, bold=True, color=GREEN)
    ws.row_dimensions[20].height = 30
    top10 = sorted([r for r in rows if r["intent"] == "transactional"],
                   key=lambda x: -int(x["monthly_volume"])) [:10]
    top10_rows = kw_rows(top10)
    write_table(ws, top10_rows, ["Keyword","Vol","CPC","Comp","Intent","Prio","Target","Notas"], GREEN, GREEN_LIGHT, start_row=21)

    set_widths(ws, [38, 10, 10, 12, 14, 10, 22, 50])
    freeze(ws, row=4, col=1)

    # =========================================================================
    # TABS DE CLUSTERS
    # =========================================================================
    cluster_configs = [
        ("money", "💰 Money Keywords", GREEN, GREEN_LIGHT),
        ("near-me", "📍 Near Me - GBP", GREEN, GREEN_LIGHT),
        ("product", "🏷️ Productos", GOLD, GOLD_LIGHT),
        ("product-niche", "⭐ Durafence (nicho)", GOLD, GOLD_LIGHT),
        ("geo-zip", "📌 ZIPs de Ocala", ORANGE, ORANGE_LIGHT),
        ("geo-county", "🗺️ Marion County", ORANGE, ORANGE_LIGHT),
        ("nearby-city", "🚗 Ciudades vecinas", ORANGE, ORANGE_LIGHT),
        ("financing", "💵 Financiamiento $0 Down", RED, RED_LIGHT),
        ("hispanic", "🇪🇸 Mercado Hispano", RED, RED_LIGHT),
        ("service", "⚡ Servicios + CTAs", BLUE, BLUE_LIGHT),
        ("blog", "📝 Blog - Content SEO", PURPLE, PURPLE_LIGHT),
    ]

    for key, title, color, bg in cluster_configs:
        items = by_cluster.get(key, [])
        if not items: continue
        ws = wb.create_sheet(title[:30])
        subtitle = f"{len(items)} keywords · {sum(int(r['monthly_volume']) for r in items):,} vol/mes total · {CLUSTER_DESC.get(key, '')}"
        build_header(ws, title, subtitle, color, bg, cols=8)
        data = kw_rows(items)
        write_table(ws, data, ["Keyword","Vol","CPC","Comp","Intent","Prio","Target Page","Notas"], color, bg)
        set_widths(ws, [38, 10, 10, 12, 14, 10, 22, 50])
        freeze(ws, row=5, col=1)

    # =========================================================================
    # TAB — NEGATIVE KEYWORDS
    # =========================================================================
    ws = wb.create_sheet("⛔ Negative KW (Ads)")
    negs = [r for r in rows if not r["target_page"]]
    subtitle = f"⚠️ COPIA Y PEGA estos keywords en Google Ads como 'negative keywords' antes de lanzar campañas. Te ahorra ~30% del presupuesto."
    build_header(ws, "⛔ Negative Keywords", subtitle, GREY, GREY_LIGHT, cols=4)
    ws.cell(row=4, column=1, value="Keyword").font = Font(size=10, bold=True, color=WHITE)
    ws.cell(row=4, column=1).fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row=4, column=2, value="Vol").font = Font(size=10, bold=True, color=WHITE)
    ws.cell(row=4, column=2).fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row=4, column=3, value="Motivo").font = Font(size=10, bold=True, color=WHITE)
    ws.cell(row=4, column=3).fill = PatternFill("solid", fgColor=GREY)
    for i in range(1,4):
        ws.cell(row=4, column=i).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=4, column=i).border = border_thin("AAAAAA")
    ws.row_dimensions[4].height = 28
    for r_idx, r in enumerate(negs, 5):
        ws.cell(row=r_idx, column=1, value=r["keyword"])
        ws.cell(row=r_idx, column=2, value=int(r["monthly_volume"]) if r["monthly_volume"].isdigit() else 0)
        ws.cell(row=r_idx, column=3, value=r["notes"])
        for c_idx in range(1, 4):
            c = ws.cell(row=r_idx, column=c_idx)
            c.font = Font(size=10, color=DARK)
            c.fill = PatternFill("solid", fgColor=GREY_LIGHT if r_idx % 2 == 0 else WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            c.border = border_thin("EEEEEE")
    set_widths(ws, [35, 12, 60])
    freeze(ws, row=5, col=1)

    # =========================================================================
    # TAB — TODAS
    # =========================================================================
    ws = wb.create_sheet("📋 Todas")
    build_header(ws, "📋 Todas las keywords",
                 f"Lista completa de {total} keywords · vol total: {total_vol:,} búsquedas/mes · ordenadas por prioridad y volumen",
                 DARK, GREY_LIGHT, cols=9)
    all_data = kw_rows(rows)
    headers = ["Keyword","Vol","CPC","Comp","Intent","Prio","Target Page","Notas"]
    write_table(ws, all_data, headers, DARK, GREY_LIGHT)
    set_widths(ws, [38, 10, 10, 12, 14, 10, 22, 50])
    freeze(ws, row=5, col=1)

    wb.save(OUT)
    return OUT, total

if __name__ == "__main__":
    out, count = build()
    print(f"✅ Excel generado: {out}")
    print(f"   {count} keywords en 13 tabs")
